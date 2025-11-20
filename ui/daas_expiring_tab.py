import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import os
import logging
from datetime import datetime, timedelta

from ui.utils import add_context_menu
from ui.dialogs import show_asset_details, create_properly_sized_dialog, format_timestamp

logger = logging.getLogger(__name__)

class DaasExpiringTab:
    def __init__(self, parent, db, app, config=None):
        self.db = db
        self.app = app
        self.config = config or {}
        self.frame = ttk.Frame(parent)
        self.setup_ui()
        
    def setup_ui(self):
        """Set up the DaaS Expiring tab UI"""
        # Frame for inventory
        inventory_frame = ttk.Frame(self.frame)
        inventory_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Controls
        control_frame = ttk.Frame(inventory_frame)
        control_frame.pack(fill="x", padx=5, pady=5)
        
        # Refresh and Export buttons
        ttk.Button(control_frame, text="Refresh", 
                command=self.refresh_inventory).pack(side="left", padx=5)
        ttk.Button(control_frame, text="Export", 
                command=self.export_inventory).pack(side="left", padx=5)
        ttk.Button(control_frame, text="Update DaaS Data", 
                command=self.update_daas_data).pack(side="left", padx=5)
                
        # Add toggle for showing deleted assets
        self.show_deleted = tk.BooleanVar(value=False)
        ttk.Checkbutton(control_frame, text="Show Deleted Assets", 
                    variable=self.show_deleted, 
                    command=self.refresh_inventory).pack(side="left", padx=20)
        
        # Filter controls frame
        filter_frame = ttk.Frame(control_frame)
        filter_frame.pack(side="right", padx=5)
        
        # Add days filter dropdown
        ttk.Label(filter_frame, text="Expiring within:").pack(side="left", padx=5)
        self.days_filter = ttk.Combobox(filter_frame, width=10, state="readonly")
        self.days_filter.pack(side="left", padx=5)
        self.days_filter.bind("<<ComboboxSelected>>", self.refresh_inventory)
        
        # Populate days dropdown
        self.days_options = ["30 days", "60 days", "90 days", "180 days", "365 days"]
        self.days_filter['values'] = self.days_options
        self.days_filter.current(2)  # Default to 90 days
        
        # Add site filter dropdown
        ttk.Label(filter_frame, text="Site:").pack(side="left", padx=5)
        self.site_filter = ttk.Combobox(filter_frame, width=10, state="readonly")
        self.site_filter.pack(side="left", padx=5)
        self.site_filter.bind("<<ComboboxSelected>>", self.refresh_inventory)
        
        # Populate site dropdown (will be filled in refresh_inventory)
        self.populate_site_dropdown()
        
        # Search frame
        search_frame = ttk.Frame(inventory_frame)
        search_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(search_frame, text="Search:").pack(side="left", padx=5)
        self.inventory_search = ttk.Entry(search_frame, width=30)
        self.inventory_search.pack(side="left", padx=5)
        self.inventory_search.bind("<KeyRelease>", self.filter_inventory)
        
        # Add context menu
        add_context_menu(self.inventory_search)
        
        # Treeview for inventory list
        columns = ("asset_id", "hostname", "serial", "model", "assigned_to", "lease_maturity", "days_left", "status", "site")
        self.inventory_tree = ttk.Treeview(inventory_frame, columns=columns, show="headings")
        
        # Define headings
        self.inventory_tree.heading("asset_id", text="Asset Tag", command=lambda: self.treeview_sort_column(self.inventory_tree, "asset_id", False))
        self.inventory_tree.heading("hostname", text="Hostname", command=lambda: self.treeview_sort_column(self.inventory_tree, "hostname", False))
        self.inventory_tree.heading("serial", text="Serial", command=lambda: self.treeview_sort_column(self.inventory_tree, "serial", False))
        self.inventory_tree.heading("model", text="Model", command=lambda: self.treeview_sort_column(self.inventory_tree, "model", False))
        self.inventory_tree.heading("assigned_to", text="Assigned To", command=lambda: self.treeview_sort_column(self.inventory_tree, "assigned_to", False))
        self.inventory_tree.heading("lease_maturity", text="Maturity Date", command=lambda: self.treeview_sort_column(self.inventory_tree, "lease_maturity", False))
        self.inventory_tree.heading("days_left", text="Days Left", command=lambda: self.treeview_sort_column(self.inventory_tree, "days_left", False))
        self.inventory_tree.heading("status", text="Status", command=lambda: self.treeview_sort_column(self.inventory_tree, "status", False))
        self.inventory_tree.heading("site", text="Site", command=lambda: self.treeview_sort_column(self.inventory_tree, "site", False))
        
        # Define columns widths
        self.inventory_tree.column("asset_id", width=100)
        self.inventory_tree.column("hostname", width=150)
        self.inventory_tree.column("serial", width=150)
        self.inventory_tree.column("model", width=200)
        self.inventory_tree.column("assigned_to", width=150)
        self.inventory_tree.column("lease_maturity", width=100)
        self.inventory_tree.column("days_left", width=80)
        self.inventory_tree.column("status", width=80)
        self.inventory_tree.column("site", width=80)
        
        # Add scrollbar
        scrollbar = ttk.Scrollbar(inventory_frame, orient="vertical", command=self.inventory_tree.yview)
        self.inventory_tree.configure(yscrollcommand=scrollbar.set)
        
        self.inventory_tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Bind double-click to view details
        self.inventory_tree.bind("<Double-1>", self.on_inventory_double_click)
        
        # Add context menu to treeview
        self.add_inventory_context_menu()
        
        # Status bar
        self.status_var = tk.StringVar(value="Ready to display expiring assets")
        status_bar = ttk.Label(inventory_frame, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        status_bar.pack(side="bottom", fill="x", pady=(5, 0))
        
        # Populate with current data
        self.refresh_inventory()

    def populate_site_dropdown(self):
        """Populate the site dropdown with available sites"""
        # Get all unique sites from the database
        conn = None
        try:
            conn = self.db.get_connection(write=False)  # Read operation
            cursor = self.db.cursor(conn)
            
            # Query to get unique sites
            if self.db.using_local:
                cursor.execute("""
                    SELECT DISTINCT h.site 
                    FROM scan_history h
                    JOIN assets a ON h.asset_id = a.asset_id
                    WHERE h.site IS NOT NULL AND h.site != '' AND a.expiry_flag_status = 1
                """)
            else:
                cursor.execute("""
                    SELECT DISTINCT h.site 
                    FROM scan_history h
                    JOIN assets a ON h.asset_id = a.asset_id
                    WHERE h.site IS NOT NULL AND h.site != '' AND a.expiry_flag_status = TRUE
                """)
            
            sites = [row[0] for row in cursor.fetchall()]
            
            # Add "All Sites" option
            sites = ["All Sites"] + sorted(sites)
            
            # Update the dropdown
            self.site_filter['values'] = sites
            
            # Set default value to "All Sites"
            self.site_filter.current(0)
            
        except Exception as e:
            logger.error(f"Error populating site dropdown: {str(e)}")
        finally:
            if conn:
                self.db.release_connection(conn)
    
    def treeview_sort_column(self, treeview, col, reverse):
        """Sort treeview content when column header is clicked"""
        l = [(treeview.set(k, col), k) for k in treeview.get_children('')]
        
        # Try to sort numerically if possible, otherwise sort as strings
        try:
            l.sort(key=lambda t: int(t[0]) if t[0].isdigit() else float('inf'), reverse=reverse)
        except ValueError:
            l.sort(reverse=reverse)
        
        # Rearrange items in sorted positions
        for index, (val, k) in enumerate(l):
            treeview.move(k, '', index)
        
        # Reverse sort next time
        treeview.heading(col, command=lambda: self.treeview_sort_column(treeview, col, not reverse))

    def add_inventory_context_menu(self):
        """Add context menu to inventory treeview"""
        menu = tk.Menu(self.inventory_tree, tearoff=0)
        
        def popup(event):
            # Select row under mouse
            iid = self.inventory_tree.identify_row(event.y)
            if iid:
                self.inventory_tree.selection_set(iid)
                menu.post(event.x_root, event.y_root)
        
        menu.add_command(label="View Details", command=self.view_selected_asset)
        menu.add_command(label="Check In", command=self.check_in_selected_asset)
        menu.add_command(label="Check Out", command=self.check_out_selected_asset)
        
        self.inventory_tree.bind("<Button-3>", popup)

    def view_selected_asset(self):
        """Show details for the selected asset"""
        selection = self.inventory_tree.selection()
        if not selection:
            return
        
        asset_id = self.inventory_tree.item(selection[0], "values")[0]
        show_asset_details(self.db, asset_id, self.config)
    
    def check_in_selected_asset(self):
        """Check in the selected asset"""
        from ui.dialogs import show_check_in_out_dialog
        
        selection = self.inventory_tree.selection()
        if not selection:
            return
        
        asset_id = self.inventory_tree.item(selection[0], "values")[0]
        asset_data = self.db.get_asset_by_id(asset_id)
        
        if asset_data:
            show_check_in_out_dialog(self.db, asset_data, default_status="in", callback=self.app.refresh_all, site_config=self.config)
    
    def check_out_selected_asset(self):
        """Check out the selected asset"""
        from ui.dialogs import show_check_in_out_dialog
        
        selection = self.inventory_tree.selection()
        if not selection:
            return
        
        asset_id = self.inventory_tree.item(selection[0], "values")[0]
        asset_data = self.db.get_asset_by_id(asset_id)
        
        if asset_data:
            show_check_in_out_dialog(self.db, asset_data, default_status="out", callback=self.app.refresh_all, site_config=self.config)
    
    def on_inventory_double_click(self, event):
        """Handle double-click on inventory item"""
        item = self.inventory_tree.identify('item', event.x, event.y)
        if item:
            asset_id = self.inventory_tree.item(item, "values")[0]
            show_asset_details(self.db, asset_id, self.config)

    def refresh_inventory(self, event=None):
        """Refresh the DaaS expiring inventory view"""
        # Clear existing items
        for item in self.inventory_tree.get_children():
            self.inventory_tree.delete(item)
        
        # Get days filter value
        days_filter = self.days_filter.get() if hasattr(self, 'days_filter') else "90 days"
        days = int(days_filter.split()[0])  # Extract the number of days
        
        # Get site filter value
        site_filter = self.site_filter.get() if hasattr(self, 'site_filter') else "All Sites"
        
        # Show deleted assets
        include_deleted = self.show_deleted.get() if hasattr(self, 'show_deleted') else False
        
        self.status_var.set(f"Loading assets expiring within {days} days...")
        self.frame.update_idletasks()
        
        # Get expiring assets
        inventory = self.db.get_expiring_assets(days, include_deleted)
        
        # Filter assets by expiry date and site
        filtered_inventory = []
        today = datetime.now().date()
        
        for item in inventory:
            maturity_date_str = item.get('lease_maturity_date')
            if not maturity_date_str:
                continue
                
            try:
                # Try parsing the maturity date
                maturity_date = None
                if isinstance(maturity_date_str, str):
                    # Try different date formats
                    for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
                        try:
                            maturity_date = datetime.strptime(maturity_date_str, fmt).date()
                            break
                        except ValueError:
                            continue
                else:
                    # If it's already a date/datetime
                    if isinstance(maturity_date_str, datetime):
                        maturity_date = maturity_date_str.date()
                    else:
                        maturity_date = maturity_date_str
                
                if not maturity_date:
                    continue
                
                # Calculate days left
                days_left = (maturity_date - today).days
                
                # Only include if within the filter range or already expired
                if days_left <= days:  # Changed to include expired items (negative days_left)
                    item['days_left'] = days_left
                    
                    # Filter by site if needed
                    if site_filter != "All Sites" and item.get('site') != site_filter:
                        continue
                        
                    filtered_inventory.append(item)
                    
            except Exception as e:
                logger.error(f"Error processing maturity date for {item.get('asset_id')}: {e}")
        
        if not filtered_inventory:
            self.status_var.set(f"No assets found expiring within {days} days" + 
                            (f" for site {site_filter}" if site_filter != "All Sites" else ""))
            return
        
        # Sort by days left (ascending)
        filtered_inventory.sort(key=lambda x: x.get('days_left', 999))
        
        # Populate treeview
        for item in filtered_inventory:
            # Determine if this is a deleted asset
            is_deleted = item.get('operational_status') == 'DELETED'
            
            # Get days left
            days_left = item.get('days_left', '')
            
            # Set urgency tags based on days left
            tags = ()
            if is_deleted:
                tags = ('deleted',)
            elif days_left < 0:
                tags = ('expired',)  # New tag for expired items
            elif days_left <= 30:
                tags = ('urgent',)
            elif days_left <= 60:
                tags = ('warning',)
            else:
                tags = ('normal',)
            
            # Format model information (combining manufacturer and model)
            model_text = f"{item.get('manufacturer', '')} {item.get('model_id', '')}"
            
            # Format status
            status = item.get('status', '').upper() if item.get('status') else ''
            
            # Format days left display
            days_left_display = f"{days_left} days"
            if days_left < 0:
                days_left_display = f"EXPIRED ({abs(days_left)} days ago)"
            
            self.inventory_tree.insert("", "end", values=(
                item.get('asset_id', ''),
                item.get('hostname', ''),
                item.get('serial_number', ''),
                model_text.strip(),
                item.get('assigned_to', ''),
                item.get('lease_maturity_date', ''),
                days_left_display,
                status,
                item.get('site', '')
            ), tags=tags)
        
        # Configure tag styling
        self.inventory_tree.tag_configure('deleted', foreground='gray', font=('', 9, 'italic'))
        self.inventory_tree.tag_configure('expired', background='#FFCDD2')  # Light red for expired
        self.inventory_tree.tag_configure('urgent', background='#FFD6D6')  # Light red
        self.inventory_tree.tag_configure('warning', background='#FFF3D6')  # Light yellow
        self.inventory_tree.tag_configure('normal', background='#F0F0F0')  # Light gray
        
        site_text = f" for site {site_filter}" if site_filter != "All Sites" else ""
        self.status_var.set(f"Found {len(filtered_inventory)} assets expiring within {days} days{site_text}")
    
    def filter_inventory(self, event=None):
        """Filter inventory based on search text"""
        search_text = self.inventory_search.get().lower()
        
        # If search text is empty, just refresh
        if not search_text:
            self.refresh_inventory()
            return
        
        # Clear existing items but keep the filtered data
        for item in self.inventory_tree.get_children():
            item_values = self.inventory_tree.item(item, 'values')
            
            # Check if search text matches any column
            if any(search_text in str(value).lower() for value in item_values):
                # Keep this item (style it to show it matches)
                self.inventory_tree.item(item, tags=self.inventory_tree.item(item, 'tags') + ('match',))
            else:
                # Hide this item
                self.inventory_tree.detach(item)
        
        # Configure match styling
        self.inventory_tree.tag_configure('match', background='#E0F0FF')  # Light blue
        
        # Update status
        visible_items = len(self.inventory_tree.get_children())
        self.status_var.set(f"Found {visible_items} matching assets")

    def export_inventory(self):
        """Export the DaaS expiring inventory to CSV"""
        from datetime import datetime
        
        # Get configurations for export
        include_deleted = self.show_deleted.get() if hasattr(self, 'show_deleted') else False
        days_filter = self.days_filter.get() if hasattr(self, 'days_filter') else "90 days"
        days = int(days_filter.split()[0])  # Extract the number of days
        site_filter = self.site_filter.get() if hasattr(self, 'site_filter') else "All Sites"
        
        # Ask user to confirm export settings
        export_options = f"Include deleted assets: {'Yes' if include_deleted else 'No'}\n"
        export_options += f"Expiring within: {days_filter}\n"
        export_options += f"Site filter: {site_filter}"
        
        result = messagebox.askyesno("Export Options", 
                                f"Confirm export with these settings?\n\n{export_options}",
                                icon='question')
        
        if not result:
            return
        
        # Get expiring assets
        inventory = self.db.get_expiring_assets(days, include_deleted)
        
        # Filter and enrich the data
        filtered_inventory = []
        today = datetime.now().date()
        
        for item in inventory:
            maturity_date_str = item.get('lease_maturity_date')
            if not maturity_date_str:
                continue
                
            try:
                # Parse the maturity date
                maturity_date = None
                if isinstance(maturity_date_str, str):
                    # Try different date formats
                    for fmt in ['%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d']:
                        try:
                            maturity_date = datetime.strptime(maturity_date_str, fmt).date()
                            break
                        except ValueError:
                            continue
                else:
                    # If it's already a date/datetime
                    if isinstance(maturity_date_str, datetime):
                        maturity_date = maturity_date_str.date()
                    else:
                        maturity_date = maturity_date_str
                
                if not maturity_date:
                    continue
                
                # Calculate days left
                days_left = (maturity_date - today).days
                
                # Only include if within the filter range or already expired
                if days_left <= days:  # Changed to include expired items
                    item['days_left'] = days_left
                    
                    # Filter by site if needed
                    if site_filter != "All Sites" and item.get('site') != site_filter:
                        continue
                        
                    filtered_inventory.append(item)
                    
            except Exception as e:
                logger.error(f"Error processing maturity date for {item.get('asset_id')}: {e}")
        
        if not filtered_inventory:
            messagebox.showinfo("No Data", f"No assets found expiring within {days} days" + 
                            (f" for site {site_filter}" if site_filter != "All Sites" else ""))
            return
        
        # Sort by days left (ascending)
        filtered_inventory.sort(key=lambda x: x.get('days_left', 999))
        
        # Ask for save location
        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"daas_expiring_{days}_days_{site_filter.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        )
        
        if not filename:
            return  # User cancelled
        
        # Write to CSV
        try:
            with open(filename, 'w') as f:
                # Define CSV headers
                headers = [
                    'asset_id', 'hostname', 'serial_number', 'manufacturer', 'model_id',
                    'assigned_to', 'lease_start_date', 'lease_maturity_date', 'days_left',
                    'status', 'site', 'operational_status'
                ]
                
                # Write headers
                f.write(','.join([f'"{h}"' for h in headers]) + '\n')
                
                # Write data rows
                for item in filtered_inventory:
                    row = []
                    for header in headers:
                        value = item.get(header, '')
                        if header == 'days_left' and header not in item:
                            value = item.get('days_left', '')
                        row.append(f'"{str(value)}"')
                    f.write(','.join(row) + '\n')
                
            site_text = f" for site {site_filter}" if site_filter != "All Sites" else ""
            messagebox.showinfo("Export Complete", 
                            f"Exported {len(filtered_inventory)} expiring assets{site_text} to:\n{filename}")
        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to export data: {str(e)}")
            logger.error(f"Export error: {e}")
    
    def update_daas_data(self):
        """Update DaaS data from the Excel/CSV file with improved sheet and column selection"""
        from tkinter import filedialog
        import os
        import threading
        import pandas as pd
        
        # Ask for file location or use default location
        default_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Decom Tracker.xlsx")
        
        if os.path.exists(default_path):
            # Ask to use default file
            use_default = messagebox.askyesno(
                "Use Default File",
                f"Found default DaaS file:\n{os.path.basename(default_path)}\n\nDo you want to use this file?",
                icon='question'
            )
            
            if use_default:
                file_path = default_path
            else:
                # Let user choose another file
                file_path = filedialog.askopenfilename(
                    filetypes=[
                        ("Excel files", "*.xlsx"),
                        ("CSV files", "*.csv"),
                        ("All files", "*.*")
                    ],
                    title="Select DaaS Data File",
                    initialdir=os.path.dirname(default_path)
                )
        else:
            # No default file, ask user to select
            file_path = filedialog.askopenfilename(
                filetypes=[
                    ("Excel files", "*.xlsx"),
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ],
                title="Select DaaS Data File"
            )
        
        if not file_path:
            return  # User cancelled
        
        self.status_var.set(f"Loading {os.path.basename(file_path)}...")
        self.frame.update_idletasks()
        
        try:
            # For Excel files, use the enhanced column selection dialog
            if file_path.lower().endswith('.xlsx'):
                # Show Excel workbook selection dialog
                excel_dialog = ExcelSheetDialog(self.frame, file_path)
                if not excel_dialog.result:
                    self.status_var.set("Update cancelled.")
                    return
                
                sheet_name = excel_dialog.result['sheet_name']
                serial_col = excel_dialog.result['serial_col']
                start_col = excel_dialog.result['start_col']
                maturity_col = excel_dialog.result['maturity_col']
                
                # Process in a separate thread
                def process_thread():
                    try:
                        # Now process with the selected sheet and columns
                        success, message = self._process_excel_file(file_path, sheet_name, serial_col, start_col, maturity_col)
                        
                        # Update UI on the main thread
                        self.frame.after(100, lambda s=success, m=message: self._handle_update_result(s, m))
                    except Exception as e:
                        # Handle exceptions in thread
                        self.frame.after(100, lambda e=e: self._handle_update_error(str(e)))
                
                threading.Thread(target=process_thread, daemon=True).start()
                
            # For CSV files, use the original approach
            elif file_path.lower().endswith('.csv'):
                # Read the file to get a preview
                df = pd.read_csv(file_path, nrows=10)  # Read just a few rows for preview
                
                # Show column selection dialog
                col_dialog = ColumnSelectionDialog(self.frame, df)
                if not col_dialog.result:
                    self.status_var.set("Update cancelled.")
                    return
                
                serial_col = col_dialog.result['serial_col']
                start_col = col_dialog.result['start_col']
                maturity_col = col_dialog.result['maturity_col']
                has_header = col_dialog.result['has_header']
                
                # Process in a separate thread
                def process_thread():
                    try:
                        # Now process with the manually selected columns
                        df = pd.read_csv(file_path, header=0 if has_header else None)
                        success, message = self._process_with_manual_columns(df, serial_col, start_col, maturity_col)
                        
                        # Update UI on the main thread
                        self.frame.after(100, lambda s=success, m=message: self._handle_update_result(s, m))
                    except Exception as e:
                        # Handle exceptions in thread
                        self.frame.after(100, lambda e=e: self._handle_update_error(str(e)))
                
                threading.Thread(target=process_thread, daemon=True).start()
                
            else:
                messagebox.showerror("Error", "Unsupported file format. Please use .xlsx or .csv", parent=self.frame)
                return
            
        except Exception as e:
            self.status_var.set(f"Error: {str(e)}")
            messagebox.showerror("Update Error", f"Failed to update DaaS data: {str(e)}")
            logger.error(f"DaaS update error: {e}")

    def _process_excel_file(self, file_path, sheet_name, serial_col, start_col, maturity_col):
        """Process the Excel file with selected sheet and columns"""
        import pandas as pd
        
        try:
            # Read the selected sheet
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Extract actual column names from the selections (removing the "A: " prefix)
            if ": " in serial_col:
                serial_col = serial_col.split(": ", 1)[1]
            
            if start_col and ": " in start_col:
                start_col = start_col.split(": ", 1)[1]
                
            if maturity_col and ": " in maturity_col:
                maturity_col = maturity_col.split(": ", 1)[1]
            
            # Process the dataframe with the real column names
            # We're passing True for has_header since Excel files are read with headers by default
            return self._process_with_manual_columns(df, serial_col, start_col, maturity_col, True)
        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            return False, f"Error processing Excel file: {str(e)}"

    def _process_with_manual_columns(self, df, serial_col, start_col, maturity_col, has_header=True):
        """Process the dataframe with manually selected columns"""
        import pandas as pd  # Add this import statement
        
        total_rows = len(df)
        updated_count = 0
        not_found_count = 0
        error_count = 0
        not_found_serials = []
        
        # Log the column information
        logger.info(f"Processing data with columns: Serial={serial_col}, Start={start_col}, Maturity={maturity_col}, Has Header={has_header}")
        
        for index, row in df.iterrows():
            try:
                # Extract data from the row using the selected columns
                serial_number = str(row[serial_col]).strip()
                
                # Skip empty rows
                if not serial_number or pd.isna(serial_number) or serial_number.lower() == 'nan':
                    continue
                
                # Parse lease dates
                lease_start_date = None
                if start_col is not None:
                    try:
                        start_date = row[start_col]
                        if not pd.isna(start_date):
                            if isinstance(start_date, str):
                                lease_start_date = start_date
                            else:
                                # Convert to string in YYYY-MM-DD format
                                try:
                                    lease_start_date = pd.to_datetime(start_date).strftime('%Y-%m-%d')
                                except Exception as e:
                                    logger.warning(f"Could not parse start date '{start_date}' for serial {serial_number}: {e}")
                    except Exception as e:
                        logger.warning(f"Error accessing start date column for row {index}: {e}")
                
                lease_maturity_date = None
                if maturity_col is not None:
                    try:
                        maturity_date = row[maturity_col]
                        if not pd.isna(maturity_date):
                            if isinstance(maturity_date, str):
                                lease_maturity_date = maturity_date
                            else:
                                # Convert to string in YYYY-MM-DD format
                                try:
                                    lease_maturity_date = pd.to_datetime(maturity_date).strftime('%Y-%m-%d')
                                except Exception as e:
                                    logger.warning(f"Could not parse maturity date '{maturity_date}' for serial {serial_number}: {e}")
                    except Exception as e:
                        logger.warning(f"Error accessing maturity date column for row {index}: {e}")
                
                # Skip if both dates are None
                if lease_start_date is None and lease_maturity_date is None:
                    continue
                
                # Find asset by serial number
                asset = self.db.get_asset_by_serial(serial_number)
                if not asset:
                    not_found_count += 1
                    not_found_serials.append(serial_number)
                    continue
                
                # Update lease info
                success = self.db.update_lease_info(
                    asset['asset_id'],
                    lease_start_date=lease_start_date,
                    lease_maturity_date=lease_maturity_date
                )
                
                if success:
                    updated_count += 1
                else:
                    error_count += 1
                    
            except Exception as e:
                logger.error(f"Error processing row {index}: {e}")
                error_count += 1
        
        # Update expiry flags for all assets
        self.db.update_all_expiry_flags()
        
        # Log results
        logger.info(f"Processed {total_rows} rows with manual column selection")
        logger.info(f"Updated: {updated_count}, Not found: {not_found_count}, Errors: {error_count}")
        
        if not_found_count > 0:
            logger.info(f"Serials not found: {not_found_serials[:10]}...")
        
        return True, f"Processed {total_rows} rows. Updated: {updated_count}, Not found: {not_found_count}, Errors: {error_count}"
    
    def _handle_update_result(self, success, message):
        """Handle the result of DaaS data update"""
        if success:
            self.status_var.set(f"Update completed: {message}")
            messagebox.showinfo("Update Complete", message)
            
            # Update expiry flags for all assets
            self.db.update_all_expiry_flags()
            
            # Refresh the view
            self.refresh_inventory()
        else:
            self.status_var.set(f"Update failed: {message}")
            messagebox.showerror("Update Failed", message)

class ColumnSelectionDialog:
    """Dialog to allow users to select the columns for lease data import"""
    def __init__(self, parent, df):
        self.result = None
        
        # Create dialog
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Select Data Columns")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Make dialog resizable
        self.dialog.resizable(True, True)
        self.dialog.minsize(700, 500)
        
        # Center the dialog
        self.dialog.update_idletasks()
        width = self.dialog.winfo_width()
        height = self.dialog.winfo_height()
        x = (parent.winfo_width() - width) // 2 + parent.winfo_x()
        y = (parent.winfo_height() - height) // 2 + parent.winfo_y()
        self.dialog.geometry(f"+{x}+{y}")
        
        # Configure dialog grid
        self.dialog.columnconfigure(0, weight=1)
        self.dialog.rowconfigure(2, weight=1)  # Preview grid row
        
        # Instructions
        ttk.Label(self.dialog, text="Please identify the columns in your file:", font=("", 12, "bold")).grid(
            row=0, column=0, padx=10, pady=(10, 5), sticky="w"
        )
        
        # Example: first few rows of the dataframe
        preview_frame = ttk.LabelFrame(self.dialog, text="Data Preview")
        preview_frame.grid(row=2, column=0, padx=10, pady=10, sticky="nsew")
        
        # Configure preview frame
        preview_frame.columnconfigure(0, weight=1)
        preview_frame.rowconfigure(0, weight=1)
        
        # Create a Text widget for the preview
        preview_text = tk.Text(preview_frame, wrap=tk.NONE, height=10, width=80)
        preview_text.grid(row=0, column=0, sticky="nsew")
        
        # Add scrollbars
        y_scrollbar = ttk.Scrollbar(preview_frame, orient=tk.VERTICAL, command=preview_text.yview)
        y_scrollbar.grid(row=0, column=1, sticky="ns")
        x_scrollbar = ttk.Scrollbar(preview_frame, orient=tk.HORIZONTAL, command=preview_text.xview)
        x_scrollbar.grid(row=1, column=0, sticky="ew")
        
        preview_text.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        
        # Insert a preview of the dataframe
        preview_text.insert(tk.END, str(df))
        preview_text.config(state=tk.DISABLED)  # Make it read-only
        
        # Header checkbox
        header_var = tk.BooleanVar(value=False)
        header_check = ttk.Checkbutton(self.dialog, text="File has header row", variable=header_var)
        header_check.grid(row=1, column=0, padx=10, pady=(5, 0), sticky="w")
        
        # Column selection controls
        selection_frame = ttk.Frame(self.dialog)
        selection_frame.grid(row=3, column=0, padx=10, pady=10, sticky="ew")
        
        # Configure 3 columns
        for i in range(3):
            selection_frame.columnconfigure(i, weight=1)
        
        # Get number of columns and create column options
        num_columns = len(df.columns)
        column_options = []
        
        if header_var.get():
            # Use column names if headers are present
            column_options = [str(col) for col in df.columns]
        else:
            # Use column indices otherwise
            column_options = [str(i) for i in range(num_columns)]
        
        # Serial Number column selection
        ttk.Label(selection_frame, text="Serial Number Column:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        serial_var = tk.StringVar()
        serial_combo = ttk.Combobox(selection_frame, textvariable=serial_var, values=column_options, state="readonly")
        serial_combo.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
        
        # Lease Start Date column selection
        ttk.Label(selection_frame, text="Lease Start Date Column:").grid(row=0, column=1, padx=5, pady=5, sticky="w")
        start_var = tk.StringVar()
        start_combo = ttk.Combobox(selection_frame, textvariable=start_var, values=column_options, state="readonly")
        start_combo.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        
        # Lease Maturity Date column selection
        ttk.Label(selection_frame, text="Lease Maturity Date Column:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        maturity_var = tk.StringVar()
        maturity_combo = ttk.Combobox(selection_frame, textvariable=maturity_var, values=column_options, state="readonly")
        maturity_combo.grid(row=1, column=2, padx=5, pady=5, sticky="ew")
        
        # Update column options when header checkbox changes
        def update_column_options():
            nonlocal column_options
            if header_var.get():
                # Use column names if headers are present
                column_options = [str(col) for col in df.columns]
            else:
                # Use column indices otherwise
                column_options = [str(i) for i in range(num_columns)]
            
            serial_combo['values'] = column_options
            start_combo['values'] = column_options
            maturity_combo['values'] = column_options
        
        header_check.config(command=update_column_options)
        
        # Try to auto-detect columns by position
        # We're assuming this structure: serial (0), start date (8), maturity date (9)
        if num_columns > 9:
            serial_combo.current(0)  # First column likely contains serial
            start_combo.current(8)   # Typically column I in Excel
            maturity_combo.current(9)  # Typically column J in Excel
        
        # Button frame
        button_frame = ttk.Frame(self.dialog)
        button_frame.grid(row=4, column=0, pady=10)
        
        def on_ok():
            serial_input = serial_var.get()
            start_input = start_var.get()
            maturity_input = maturity_var.get()
            
            if not serial_input:
                messagebox.showerror("Error", "Please select Serial Number column", parent=self.dialog)
                return
                
            # Convert column selections from strings to proper indices or names
            try:
                serial_col = int(serial_input) if serial_input.isdigit() else serial_input
                start_col = int(start_input) if start_input and start_input.isdigit() else start_input
                maturity_col = int(maturity_input) if maturity_input and maturity_input.isdigit() else maturity_input
                
                # For empty start/maturity selections, use None
                if not start_input:
                    start_col = None
                if not maturity_input:
                    maturity_col = None
                
                self.result = {
                    'serial_col': serial_col,
                    'start_col': start_col,
                    'maturity_col': maturity_col,
                    'has_header': header_var.get()
                }
                self.dialog.destroy()
            except ValueError as e:
                messagebox.showerror("Error", f"Invalid column selection: {str(e)}", parent=self.dialog)
        
        ttk.Button(button_frame, text="OK", command=on_ok).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.dialog.destroy).pack(side=tk.LEFT, padx=5)
        
        # Wait for dialog to close
        parent.wait_window(self.dialog)

class ExcelSheetDialog:
    """Dialog to select a sheet and columns from an Excel workbook"""
    def __init__(self, parent, file_path):
        import pandas as pd
        import openpyxl
        
        self.result = None
        
        # Create dialog
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Select Excel Sheet and Columns")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Make dialog resizable
        self.dialog.resizable(True, True)
        self.dialog.minsize(900, 600)
        
        # Center the dialog
        self.dialog.update_idletasks()
        width = self.dialog.winfo_width()
        height = self.dialog.winfo_height()
        x = (parent.winfo_width() - width) // 2 + parent.winfo_x()
        y = (parent.winfo_height() - height) // 2 + parent.winfo_y()
        self.dialog.geometry(f"+{x}+{y}")
        
        # Configure dialog grid
        self.dialog.columnconfigure(0, weight=1)
        self.dialog.rowconfigure(3, weight=1)  # Preview grid row
        
        # Instructions
        ttk.Label(self.dialog, text="Step 1: Select a worksheet from the Excel file", font=("", 12, "bold")).grid(
            row=0, column=0, padx=10, pady=(10, 5), sticky="w"
        )
        
        # Sheet selection
        sheet_frame = ttk.Frame(self.dialog)
        sheet_frame.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="ew")
        
        ttk.Label(sheet_frame, text="Worksheet:").pack(side=tk.LEFT, padx=(0, 5))
        
        # Get sheet names from the workbook
        self.workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = self.workbook.sheetnames
        
        # Store file path
        self.file_path = file_path
        
        # Sheet selection combobox
        self.sheet_var = tk.StringVar()
        sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.sheet_var, values=sheet_names, state="readonly", width=40)
        sheet_combo.pack(side=tk.LEFT, padx=5)
        
        if sheet_names:
            sheet_combo.current(0)  # Select first sheet by default
        
        ttk.Button(sheet_frame, text="Load Preview", command=self.load_sheet_preview).pack(side=tk.LEFT, padx=5)
        
        # Column selection instructions
        ttk.Label(self.dialog, text="Step 2: Identify the columns in your worksheet:", font=("", 12, "bold")).grid(
            row=2, column=0, padx=10, pady=(10, 5), sticky="w"
        )
        
        # Preview frame
        self.preview_frame = ttk.LabelFrame(self.dialog, text="Data Preview")
        self.preview_frame.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")
        
        # Configure preview frame
        self.preview_frame.columnconfigure(0, weight=1)
        self.preview_frame.rowconfigure(0, weight=1)
        
        # Create a Text widget for the preview
        self.preview_text = tk.Text(self.preview_frame, wrap=tk.NONE, height=15, width=90)
        self.preview_text.grid(row=0, column=0, sticky="nsew")
        
        # Add scrollbars
        y_scrollbar = ttk.Scrollbar(self.preview_frame, orient=tk.VERTICAL, command=self.preview_text.yview)
        y_scrollbar.grid(row=0, column=1, sticky="ns")
        x_scrollbar = ttk.Scrollbar(self.preview_frame, orient=tk.HORIZONTAL, command=self.preview_text.xview)
        x_scrollbar.grid(row=1, column=0, sticky="ew")
        
        self.preview_text.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        
        # Column selection controls
        selection_frame = ttk.Frame(self.dialog)
        selection_frame.grid(row=4, column=0, padx=10, pady=10, sticky="ew")
        
        # Configure grid
        for i in range(3):
            selection_frame.columnconfigure(i, weight=1)
        
        # Serial Number column selection
        ttk.Label(selection_frame, text="Serial Number Column:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.serial_var = tk.StringVar()
        self.serial_combo = ttk.Combobox(selection_frame, textvariable=self.serial_var, state="readonly", width=30)
        self.serial_combo.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
        
        # Lease Start Date column selection
        ttk.Label(selection_frame, text="Lease Start Date Column:").grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.start_var = tk.StringVar()
        self.start_combo = ttk.Combobox(selection_frame, textvariable=self.start_var, state="readonly", width=30)
        self.start_combo.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        
        # Lease Maturity Date column selection
        ttk.Label(selection_frame, text="Lease Maturity Date Column:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.maturity_var = tk.StringVar()
        self.maturity_combo = ttk.Combobox(selection_frame, textvariable=self.maturity_var, state="readonly", width=30)
        self.maturity_combo.grid(row=1, column=2, padx=5, pady=5, sticky="ew")
        
        # Sample data display (to help identify columns)
        self.sample_frame = ttk.LabelFrame(self.dialog, text="Column Sample Data")
        self.sample_frame.grid(row=5, column=0, padx=10, pady=10, sticky="ew")
        
        sample_grid = ttk.Frame(self.sample_frame)
        sample_grid.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Configure grid columns
        for i in range(3):
            sample_grid.columnconfigure(i, weight=1)
        
        # Labels for sample data
        self.serial_sample = ttk.Label(sample_grid, text="", wraplength=250)
        self.serial_sample.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        
        self.start_sample = ttk.Label(sample_grid, text="", wraplength=250)
        self.start_sample.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        self.maturity_sample = ttk.Label(sample_grid, text="", wraplength=250)
        self.maturity_sample.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        
        # Update sample data when column selections change
        self.serial_combo.bind("<<ComboboxSelected>>", self.update_sample_data)
        self.start_combo.bind("<<ComboboxSelected>>", self.update_sample_data)
        self.maturity_combo.bind("<<ComboboxSelected>>", self.update_sample_data)
        
        # Button frame
        button_frame = ttk.Frame(self.dialog)
        button_frame.grid(row=6, column=0, pady=10)
        
        def on_ok():
            sheet_name = self.sheet_var.get()
            serial_col = self.serial_var.get()
            start_col = self.start_var.get()
            maturity_col = self.maturity_var.get()
            
            if not sheet_name:
                messagebox.showerror("Error", "Please select a worksheet", parent=self.dialog)
                return
                
            if not serial_col:
                messagebox.showerror("Error", "Please select Serial Number column", parent=self.dialog)
                return
            
            # Start and maturity date columns are optional, but at least one should be provided
            if not start_col and not maturity_col:
                if not messagebox.askyesno("Warning", 
                                       "No date columns selected. Continue anyway?", 
                                       parent=self.dialog):
                    return
            
            self.result = {
                'sheet_name': sheet_name,
                'serial_col': serial_col,
                'start_col': start_col if start_col else None,
                'maturity_col': maturity_col if maturity_col else None
            }
            self.dialog.destroy()
        
        ttk.Button(button_frame, text="OK", command=on_ok).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.dialog.destroy).pack(side=tk.LEFT, padx=5)
        
        # Load preview of the first sheet
        if sheet_names:
            self.load_sheet_preview()
        
        # Wait for dialog to close
        parent.wait_window(self.dialog)
    
    def load_sheet_preview(self):
        """Load preview of the selected sheet"""
        import pandas as pd
        
        sheet_name = self.sheet_var.get()
        if not sheet_name:
            return
        
        try:
            # Read the selected sheet
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, nrows=15)
            
            # Clear preview text
            self.preview_text.config(state=tk.NORMAL)
            self.preview_text.delete(1.0, tk.END)
            
            # Format and display preview
            preview_text = f"Preview of sheet '{sheet_name}':\n\n"
            preview_text += df.to_string(index=True)
            self.preview_text.insert(tk.END, preview_text)
            self.preview_text.config(state=tk.DISABLED)
            
            # Update column selection dropdowns
            column_options = []
            
            # Add column letters (Excel style)
            for i, col in enumerate(df.columns):
                col_letter = self.get_column_letter(i)
                display_name = f"{col_letter}: {col}"
                column_options.append(display_name)
            
            # Update comboboxes
            self.serial_combo['values'] = column_options
            self.start_combo['values'] = column_options
            self.maturity_combo['values'] = column_options
            
            # Clear selections
            self.serial_combo.set('')
            self.start_combo.set('')
            self.maturity_combo.set('')
            
            # Try to auto-detect columns
            self.auto_detect_columns(df)
            
            # Store dataframe for sample data
            self.df = df
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheet: {str(e)}", parent=self.dialog)
            logger.error(f"Error loading Excel sheet: {e}")
    
    def get_column_letter(self, col_index):
        """Convert 0-based column index to Excel-style column letter"""
        result = ""
        while col_index >= 0:
            col_index, remainder = divmod(col_index, 26)
            result = chr(65 + remainder) + result
            col_index -= 1
        return result
    
    def auto_detect_columns(self, df):
        """Try to auto-detect relevant columns"""
        # Look for keywords in column names
        serial_keywords = ['serial', 'sn', 'device', 'asset']
        start_keywords = ['start', 'begin', 'lease start']
        maturity_keywords = ['end', 'maturity', 'expiry', 'expire', 'term']
        
        # For each column, check if any keyword is in the column name
        for i, col in enumerate(df.columns):
            col_letter = self.get_column_letter(i)
            col_str = str(col).lower()
            
            # Check for serial number column
            if any(kw in col_str for kw in serial_keywords):
                self.serial_combo.set(f"{col_letter}: {col}")
            
            # Check for lease start date column
            if any(kw in col_str for kw in start_keywords):
                self.start_combo.set(f"{col_letter}: {col}")
            
            # Check for lease maturity date column
            if any(kw in col_str for kw in maturity_keywords):
                self.maturity_combo.set(f"{col_letter}: {col}")
        
        # If no matches found, try excel standard columns for our expected data
        if not self.serial_var.get() and len(df.columns) > 0:
            self.serial_combo.current(0)  # First column (A) for serial
        
        if not self.start_var.get() and len(df.columns) > 8:
            self.start_combo.current(8)  # Column I for start date
            
        if not self.maturity_var.get() and len(df.columns) > 9:
            self.maturity_combo.current(9)  # Column J for maturity date
        
        # Update sample data
        self.update_sample_data()
    
    def update_sample_data(self, event=None):
        """Update sample data display based on selected columns"""
        if not hasattr(self, 'df') or self.df is None:
            return
            
        try:
            # Get selected columns
            serial_selection = self.serial_var.get()
            start_selection = self.start_var.get()
            maturity_selection = self.maturity_var.get()
            
            # Extract data samples
            if serial_selection:
                try:
                    col_name = serial_selection.split(": ", 1)[1] if ": " in serial_selection else serial_selection
                    sample_data = self.df[col_name].dropna().head(5).tolist()
                    self.serial_sample.config(text=f"Serial Number Samples:\n{', '.join(str(x) for x in sample_data)}")
                except Exception as e:
                    self.serial_sample.config(text=f"Error loading samples: {str(e)}")
            else:
                self.serial_sample.config(text="")
                
            if start_selection:
                try:
                    col_name = start_selection.split(": ", 1)[1] if ": " in start_selection else start_selection
                    sample_data = self.df[col_name].dropna().head(5).tolist()
                    self.start_sample.config(text=f"Start Date Samples:\n{', '.join(str(x) for x in sample_data)}")
                except Exception as e:
                    self.start_sample.config(text=f"Error loading samples: {str(e)}")
            else:
                self.start_sample.config(text="")
                
            if maturity_selection:
                try:
                    col_name = maturity_selection.split(": ", 1)[1] if ": " in maturity_selection else maturity_selection
                    sample_data = self.df[col_name].dropna().head(5).tolist()
                    self.maturity_sample.config(text=f"Maturity Date Samples:\n{', '.join(str(x) for x in sample_data)}")
                except Exception as e:
                    self.maturity_sample.config(text=f"Error loading samples: {str(e)}")
            else:
                self.maturity_sample.config(text="")
                
        except Exception as e:
            logger.error(f"Error updating sample data: {e}")
            # Don't re-raise the exception, just log it