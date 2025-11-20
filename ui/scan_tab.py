import tkinter as tk
from tkinter import ttk, messagebox
import os
import threading
import logging
import datetime  # Add this import for direct datetime formatting
from ui.utils import add_context_menu
from ui.dialogs import show_manual_entry_form, show_check_in_out_dialog, create_properly_sized_dialog, show_audit_dialog, show_bulk_checkout_dialog
from services.servicenow import scrape_servicenow

logger = logging.getLogger(__name__)

class ScanTab:
    def __init__(self, parent, db, app, config=None):
        self.db = db
        self.app = app
        self.config = config or {}
        self.frame = ttk.Frame(parent)
        self.setup_ui()
        
    def setup_ui(self):
        # Frame for scan input
        input_frame = ttk.LabelFrame(self.frame, text="Asset Input")
        input_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Site information display
        site = self.config.get('site', 'Unknown')
        site_frame = ttk.Frame(input_frame)
        site_frame.grid(row=0, column=0, columnspan=4, sticky="ew", padx=5, pady=5)
        
        site_label = ttk.Label(site_frame, text=f"Site: {site}", font=("", 10, "bold"))
        site_label.pack(side="left", padx=5)
        
        # Radio buttons for input type
        self.input_type = tk.StringVar(value="asset")
        radio_frame = ttk.Frame(input_frame)
        radio_frame.grid(row=1, column=0, columnspan=4, sticky="ew", padx=5, pady=5)
        
        ttk.Radiobutton(radio_frame, text="Asset Tag", variable=self.input_type, 
                        value="asset").pack(side="left", padx=5)
        ttk.Radiobutton(radio_frame, text="Serial Number", variable=self.input_type, 
                        value="serial").pack(side="left", padx=5)
        
        # Entry field for scanning/typing and Look Up button inline
        ttk.Label(input_frame, text="Scan or Enter:").grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.input_entry = ttk.Entry(input_frame, width=30)
        self.input_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")
        self.input_entry.focus()  # Auto-focus for scanner
        
        # Add Look Up Asset button inline with entry field
        ttk.Button(input_frame, text="Look Up Asset", width=15,
                command=self.process_asset_lookup).grid(row=2, column=2, padx=5, pady=5)
        
        # Add context menu for input entry
        add_context_menu(self.input_entry)
        
        # Bind Enter key to processing the input
        self.input_entry.bind("<Return>", lambda e: self.process_asset_lookup())
        
        # Action buttons row
        button_frame = ttk.Frame(input_frame)
        button_frame.grid(row=3, column=0, columnspan=4, pady=10)
        
        ttk.Button(button_frame, text="Manually Add", width=15,
                command=self.show_manual_entry_form).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Bulk Check-In", width=15,
                command=self.open_bulk_check_in).pack(side="left", padx=5)
        ttk.Button(button_frame, text="Bulk Check-Out", width=15,
                command=self.open_bulk_check_out).pack(side="left", padx=5)
        
        # Add DaaS Update button
        ttk.Button(button_frame, text="DaaS Update", width=15,
                command=self.update_daas_data).pack(side="left", padx=5)
        
        # Results display
        ttk.Label(input_frame, text="Asset Details:").grid(row=4, column=0, padx=5, pady=5, sticky="nw")
        self.details_text = tk.Text(input_frame, width=60, height=15, wrap="word")
        self.details_text.grid(row=4, column=1, columnspan=3, padx=5, pady=5, sticky="nsew")
        self.details_text.config(state="disabled")  # Read-only initially
        
        # Add context menu for details
        add_context_menu(self.details_text)
        
        # Status bar for showing scraping progress
        self.status_var = tk.StringVar(value="Ready")
        ttk.Label(input_frame, textvariable=self.status_var).grid(row=5, column=0, columnspan=4, pady=5)
        
        # Configure column weights to make the entry field expandable
        input_frame.columnconfigure(1, weight=1)
        input_frame.rowconfigure(4, weight=1)  # Make the details text expandable vertically

    def open_bulk_check_in(self):
        """Opens the Bulk Check-In dialog"""
        from ui.dialogs import show_bulk_checkin_dialog
        show_bulk_checkin_dialog(self.frame, self.db, self.config)
        
    def open_bulk_check_out(self):
        from ui.dialogs import show_bulk_checkout_dialog
        show_bulk_checkout_dialog(self.frame, self.db, self.config)
    
    def process_asset_lookup(self):
        """Process the asset lookup from barcode scanner or manual entry"""
        identifier = self.input_entry.get().strip()
        if not identifier:
            messagebox.showwarning("Input Required", "Please scan or enter an asset tag or serial number")
            return
            
        # Determine if this is an asset tag or serial number
        is_asset = self.input_type.get() == "asset"
        
        # Update status
        self.status_var.set(f"Looking up {'asset' if is_asset else 'serial'}: {identifier}")
        
        # Normalize case for asset tags and serial numbers
        # Asset tags typically uppercase, serials could be mixed
        if is_asset:
            identifier = identifier.upper()  # Convert asset tags to uppercase
        
        # Check if asset already exists in database
        asset = None
        if is_asset:
            asset = self.db.get_asset_by_id(identifier)
        else:
            # Try case-insensitive search for serial number
            asset = self.db.get_asset_by_serial(identifier)
            
            # If not found, try again with uppercase and lowercase versions
            if not asset:
                asset = self.db.get_asset_by_serial(identifier.upper())
            if not asset:
                asset = self.db.get_asset_by_serial(identifier.lower())
        
        if asset:
            logger.info(f"Asset found in database: {identifier}")
            
            # Check the current status (in/out) of the asset
            current_status = self.db.get_asset_current_status(asset['asset_id'])
            
            # Check if the asset is flagged
            flag_status = self.db.get_flag_status(asset['asset_id'])
            is_flagged = flag_status and flag_status.get('flag_status', False)
            
            # Display asset details with status information
            self.display_asset_details(asset, current_status)
            
            # If the asset is flagged, show a notification
            if is_flagged:
                flag_notes = flag_status.get('flag_notes', 'No reason provided')
                flag_tech = flag_status.get('flag_tech', 'Unknown')
                
                # Get the raw timestamp
                raw_timestamp = flag_status.get('flag_timestamp', '')
                
                # Manually format the timestamp with a fixed format
                flag_timestamp = "Unknown"
                if raw_timestamp:
                    try:
                        # Try to extract just the date and time without seconds or timezone
                        import re
                        # Match pattern like 2025-04-09 12:16:34.256759-05:00
                        pattern = re.compile(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}):\d{2}\.\d+-(\d{2}):.*')
                        match = pattern.match(str(raw_timestamp))
                        if match:
                            # Format as YYYY-MM-DD HH:MM -Z
                            flag_timestamp = f"{match.group(1)} -{match.group(2)}"
                        else:
                            # Try another approach - split and take just what we need
                            timestamp_parts = str(raw_timestamp).split('.')
                            if len(timestamp_parts) > 0:
                                date_time = timestamp_parts[0]
                                # Remove the seconds
                                date_time = re.sub(r'(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}):\d{2}', r'\1', date_time)
                                
                                # Get timezone if present
                                timezone = ""
                                if len(timestamp_parts) > 1 and '-' in timestamp_parts[1]:
                                    tz_match = re.search(r'-(\d{2}):', timestamp_parts[1])
                                    if tz_match:
                                        timezone = f" -{tz_match.group(1)}"
                                flag_timestamp = f"{date_time}{timezone}"
                            else:
                                flag_timestamp = str(raw_timestamp)
                    except Exception as e:
                        logger.error(f"Error formatting timestamp: {str(e)}")
                        flag_timestamp = str(raw_timestamp)
                
                # Use properly sized dialog instead of Toplevel
                from ui.dialogs import create_properly_sized_dialog
                flag_window = create_properly_sized_dialog("⚠️ FLAGGED ASSET FOUND ⚠️", 500, 400)
                
                # Configure the window background
                flag_window.configure(background="#FF9800")  # Orange background
                
                # Main content frame with padding
                content_frame = ttk.Frame(flag_window)
                content_frame.grid(row=0, column=0, sticky="nsew", padx=20, pady=20)
                
                # Configure content frame to expand properly
                flag_window.columnconfigure(0, weight=1)
                flag_window.rowconfigure(0, weight=1)
                content_frame.columnconfigure(0, weight=1)
                
                # Big flag icon
                flag_canvas = tk.Canvas(content_frame, width=50, height=50, bg="#FF9800", highlightthickness=0)
                flag_canvas.pack(pady=10)
                
                # Draw a bigger flag shape
                flag_canvas.create_rectangle(10, 5, 15, 45, fill="#FF9800", outline="black", width=2)
                flag_canvas.create_polygon(15, 5, 40, 15, 15, 25, fill="#FF9800", outline="black", width=2)
                
                ttk.Label(content_frame, text="FLAGGED ASSET DETECTED", font=("", 16, "bold")).pack(pady=5)
                ttk.Label(content_frame, text=f"Asset ID: {asset['asset_id']}", font=("", 12)).pack(pady=2)
                
                flag_frame = ttk.Frame(content_frame)
                flag_frame.pack(fill="x", pady=5)
                
                ttk.Label(flag_frame, text="Flagged by:", font=("", 10, "bold")).grid(row=0, column=0, sticky="e", padx=5)
                ttk.Label(flag_frame, text=flag_tech).grid(row=0, column=1, sticky="w", padx=5)
                
                ttk.Label(flag_frame, text="Date:", font=("", 10, "bold")).grid(row=1, column=0, sticky="e", padx=5)
                ttk.Label(flag_frame, text=flag_timestamp).grid(row=1, column=1, sticky="w", padx=5)
                
                ttk.Label(content_frame, text="Reason:", font=("", 10, "bold")).pack(anchor="w", padx=5)
                reason_text = tk.Text(content_frame, height=3, wrap="word")
                reason_text.pack(fill="x", padx=5, pady=5)
                reason_text.insert("1.0", flag_notes)
                reason_text.config(state="disabled")
                
                # Button at fixed position at bottom
                button_frame = ttk.Frame(flag_window)
                button_frame.grid(row=1, column=0, sticky="ew", padx=20, pady=20)
                
                ttk.Button(button_frame, text="Acknowledge", 
                        command=lambda: (flag_window.destroy(), 
                                        show_check_in_out_dialog(self.db, asset, 
                                                                default_status="in" if current_status.get('status') == 'out' else "out", 
                                                                current_status=current_status, 
                                                                callback=self.app.refresh_all,
                                                                site_config=self.config))).pack(pady=5)
                
                # Don't show the regular check-in/out dialog until flag is acknowledged
                return
            
            # Show check-in/out dialog with the appropriate default action
            default_status = "in" if current_status.get('status') == 'out' else "out"
            show_check_in_out_dialog(self.db, asset, default_status=default_status, 
                                    current_status=current_status, callback=self.app.refresh_all,
                                    site_config=self.config)
            
            # Clear input and update status
            self.input_entry.delete(0, "end")
            self.status_var.set("Ready")
            return
        
        # Asset not found in database, attempt to scrape from ServiceNow
        self.status_var.set(f"Fetching data from ServiceNow for {identifier}...")
        logger.info(f"Asset not found in database, attempting to scrape from ServiceNow: {identifier}")
        
        # Define scrape function for threading
        def scrape_thread():
            try:
                asset_data = scrape_servicenow(identifier, is_asset)
                # Fix: Use a lambda with a captured parameter instead of referring to a free variable
                self.frame.after(100, lambda data=asset_data: self.handle_scrape_result(data))
            except Exception as ex:
                # Fix: Capture the exception in a local variable
                error_message = str(ex)
                logger.error(f"Error scraping ServiceNow: {error_message}")
                # Use a local variable to avoid the free variable error
                self.frame.after(100, lambda msg=error_message: self.status_var.set(f"Error: {msg}"))
        
        # Start scraping in a separate thread
        threading.Thread(target=scrape_thread).start()

    def handle_scrape_result(self, asset_data):
        """Handle the result of scraping"""
        if not asset_data:
            self.status_var.set("Failed to retrieve asset data")
            messagebox.showerror("Error", "Could not retrieve asset data from ServiceNow")
            return
        
        # Ensure we have an asset_tag
        if 'asset_tag' not in asset_data and 'asset_id' in asset_data:
            asset_data['asset_tag'] = asset_data['asset_id']
        
        logger.info(f"Processing scraped asset data: {asset_data.get('asset_tag', 'Unknown')}")
        
        # Update the database
        success = self.db.update_asset(asset_data)
        
        if not success:
            logger.error("Failed to update asset in database")
            messagebox.showerror("Error", "Failed to save asset data to database")
            return
        
        # Local Save Success
        if success and self.app.db.using_local and self.app.db.pending_sync:
            self.app.show_local_save_notification()

        # Automatically check in the asset for new entries
        asset_id = asset_data.get('asset_tag', '')
        tech_name = os.getenv('USERNAME', '')
        site = self.config.get('site', 'Unknown')
        
        # Automatically check the asset in with a system-generated note
        self.db.record_scan(asset_id, "in", tech_name, "Initial automatic check-in", site)
        logger.info(f"Automatically checked in new asset: {asset_id} at site: {site}")
        
        # Display asset details (now with current status after check-in)
        current_status = self.db.get_asset_current_status(asset_id)
        self.display_asset_details(asset_data, current_status)
        
        # Update status
        self.status_var.set("Ready")
        
        logger.info("Showing check in/out dialog")
        
        # Clear input field
        self.input_entry.delete(0, "end")
        
        # Show check-in/out dialog (still showing it, but now the asset is already checked in)
        show_check_in_out_dialog(self.db, asset_data, default_status="out", 
                                current_status=current_status, callback=self.app.refresh_all, 
                                site_config=self.config)
    
    def display_asset_details(self, asset_data, current_status=None):
        """Display asset details in the main window"""
        self.details_text.config(state="normal")
        self.details_text.delete("1.0", "end")
        
        # Basic asset details
        details = f"""Asset Tag: {asset_data.get('asset_tag', '')}
Serial Number: {asset_data.get('serial_number', '')}
Hostname: {asset_data.get('hostname', '')}
Make/Model: {asset_data.get('manufacturer', '')} {asset_data.get('model_description', '')}
Assigned To: {asset_data.get('assigned_to', '')}
Location: {asset_data.get('location', '')}
OS: {asset_data.get('os', '')} {asset_data.get('os_version', '')}
Warranty Expires: {asset_data.get('warranty_expiration', '')}
"""
        
        # Add current status information if available
        if current_status:
            status_text = current_status.get('status', 'unknown')
            timestamp = current_status.get('timestamp', '')
            tech_name = current_status.get('tech_name', '')
            site = current_status.get('site', '')
            
            details += f"""
Current Status: {status_text.upper()}
Last Action: {timestamp}
Technician: {tech_name}
"""
            if site:
                details += f"Site: {site}\n"
        
        # Add site information from config
        if self.config and self.config.get('site'):
            details += f"\nCurrent Site: {self.config.get('site')}\n"
        
        # Add comments
        details += f"\nComments: {asset_data.get('comments', '')}"
        
        self.details_text.insert("1.0", details)
        self.details_text.config(state="disabled")
    
    def show_manual_entry_form(self, identifier=None):
        """Show manual data entry form"""
        identifier = self.input_entry.get().strip() if identifier is None else identifier
        
        # Call the form and handle the result
        result = show_manual_entry_form(self.db, identifier, callback=self.app.refresh_all)
        
        if result:
            # Display asset details
            self.display_asset_details(result)
            
            # Clear input
            self.input_entry.delete(0, "end")

    def update_daas_data(self):
        """Update DaaS data from the Excel/CSV file with improved sheet and column selection"""
        from tkinter import filedialog
        import os
        import threading
        import pandas as pd
        
        # Ask for file location or use default location
        default_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Decom Tracker.xlsx")
        
        if os.path.exists(default_path):
            # Ask to use default file
            use_default = messagebox.askyesno(
                "Use Default File",
                f"Found default DaaS file:\n{os.path.basename(default_path)}\n\nDo you want to use this file?",
                icon='question'
            )
            
            if use_default:
                file_path = default_path
            else:
                # Let user choose another file
                file_path = filedialog.askopenfilename(
                    filetypes=[
                        ("Excel files", "*.xlsx"),
                        ("CSV files", "*.csv"),
                        ("All files", "*.*")
                    ],
                    title="Select DaaS Data File",
                    initialdir=os.path.dirname(default_path)
                )
        else:
            # No default file, ask user to select
            file_path = filedialog.askopenfilename(
                filetypes=[
                    ("Excel files", "*.xlsx"),
                    ("CSV files", "*.csv"),
                    ("All files", "*.*")
                ],
                title="Select DaaS Data File"
            )
        
        if not file_path:
            return  # User cancelled
        
        self.status_var.set(f"Loading {os.path.basename(file_path)}...")
        self.frame.update_idletasks()
        
        try:
            # For Excel files, use the enhanced column selection dialog
            if file_path.lower().endswith('.xlsx'):
                # Show Excel workbook selection dialog
                excel_dialog = ExcelSheetDialog(self.frame, file_path)
                if not excel_dialog.result:
                    self.status_var.set("Update cancelled.")
                    return
                
                sheet_name = excel_dialog.result['sheet_name']
                serial_col = excel_dialog.result['serial_col']
                start_col = excel_dialog.result['start_col']
                maturity_col = excel_dialog.result['maturity_col']
                
                # Process in a separate thread
                def process_thread():
                    try:
                        # Now process with the selected sheet and columns
                        success, message = self._process_excel_file(file_path, sheet_name, serial_col, start_col, maturity_col)
                        
                        # Update UI on the main thread
                        self.frame.after(100, lambda s=success, m=message: self._handle_update_result(s, m))
                    except Exception as e:
                        # Handle exceptions in thread
                        self.frame.after(100, lambda e=e: self._handle_update_error(str(e)))
                
                threading.Thread(target=process_thread, daemon=True).start()
                
            # For CSV files, use the original approach
            elif file_path.lower().endswith('.csv'):
                # Read the file to get a preview
                df = pd.read_csv(file_path, nrows=10)  # Read just a few rows for preview
                
                # Show column selection dialog
                col_dialog = ColumnSelectionDialog(self.frame, df)
                if not col_dialog.result:
                    self.status_var.set("Update cancelled.")
                    return
                
                serial_col = col_dialog.result['serial_col']
                start_col = col_dialog.result['start_col']
                maturity_col = col_dialog.result['maturity_col']
                has_header = col_dialog.result['has_header']
                
                # Process in a separate thread
                def process_thread():
                    try:
                        # Now process with the manually selected columns
                        df = pd.read_csv(file_path, header=0 if has_header else None)
                        success, message = self._process_with_manual_columns(df, serial_col, start_col, maturity_col)
                        
                        # Update UI on the main thread
                        self.frame.after(100, lambda s=success, m=message: self._handle_update_result(s, m))
                    except Exception as e:
                        # Handle exceptions in thread
                        self.frame.after(100, lambda e=e: self._handle_update_error(str(e)))
                
                threading.Thread(target=process_thread, daemon=True).start()
                
            else:
                messagebox.showerror("Error", "Unsupported file format. Please use .xlsx or .csv", parent=self.frame)
                return
            
        except Exception as e:
            self.status_var.set(f"Error: {str(e)}")
            messagebox.showerror("Update Error", f"Failed to update DaaS data: {str(e)}")
            logger.error(f"DaaS update error: {e}")

    def _handle_update_result(self, success, message):
        """Handle the result of DaaS data update"""
        if success:
            self.status_var.set(f"Update completed: {message}")
            messagebox.showinfo("Update Complete", message)
            
            # Update expiry flags for all assets
            self.db.update_all_expiry_flags()
            
            # Refresh the main app
            self.app.refresh_all()
            
            # Suggest opening the DaaS Expiring tab
            open_tab = messagebox.askyesno(
                "View Expiring Assets",
                "Do you want to view assets with expiring leases?",
                icon='question'
            )
            
            if open_tab:
                # Find the DaaS Expiring tab index
                try:
                    notebook = self.app.notebook
                    for i in range(notebook.index("end")):
                        if "DaaS" in notebook.tab(i, "text"):
                            notebook.select(i)
                            break
                except Exception as e:
                    logger.error(f"Error switching to DaaS tab: {e}")
        else:
            self.status_var.set(f"Update failed: {message}")
            messagebox.showerror("Update Failed", message)

    def _handle_update_error(self, error_message):
        """Handle errors from the update thread"""
        self.status_var.set(f"Error: {error_message}")
        messagebox.showerror("Update Error", f"Failed to update DaaS data:\n{error_message}")
        logger.error(f"DaaS update thread error: {error_message}")

    def _process_excel_file(self, file_path, sheet_name, serial_col, start_col, maturity_col):
        """Process the Excel file with selected sheet and columns"""
        import pandas as pd
        
        try:
            # Read the selected sheet
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            
            # Extract actual column names from the selections (removing the "A: " prefix)
            if ": " in serial_col:
                serial_col = serial_col.split(": ", 1)[1]
            
            if start_col and ": " in start_col:
                start_col = start_col.split(": ", 1)[1]
                
            if maturity_col and ": " in maturity_col:
                maturity_col = maturity_col.split(": ", 1)[1]
            
            # Process the dataframe with the real column names
            # We're passing True for has_header since Excel files are read with headers by default
            return self._process_with_manual_columns(df, serial_col, start_col, maturity_col, True)
        except Exception as e:
            logger.error(f"Error processing Excel file: {e}")
            return False, f"Error processing Excel file: {str(e)}"
        
    def _process_with_manual_columns(self, df, serial_col, start_col, maturity_col, has_header=True):
        """Process the dataframe with manually selected columns"""
        import pandas as pd 
        
        total_rows = len(df)
        updated_count = 0
        not_found_count = 0
        error_count = 0
        not_found_serials = []
        
        # Log the column information
        logger.info(f"Processing data with columns: Serial={serial_col}, Start={start_col}, Maturity={maturity_col}, Has Header={has_header}")
        
        for index, row in df.iterrows():
            try:
                # Extract data from the row using the selected columns
                serial_number = str(row[serial_col]).strip()
                
                # Skip empty rows
                if not serial_number or pd.isna(serial_number) or serial_number.lower() == 'nan':
                    continue
                
                # Parse lease dates
                lease_start_date = None
                if start_col is not None:
                    try:
                        start_date = row[start_col]
                        if not pd.isna(start_date):
                            if isinstance(start_date, str):
                                lease_start_date = start_date
                            else:
                                # Convert to string in YYYY-MM-DD format
                                try:
                                    lease_start_date = pd.to_datetime(start_date).strftime('%Y-%m-%d')
                                except Exception as e:
                                    logger.warning(f"Could not parse start date '{start_date}' for serial {serial_number}: {e}")
                    except Exception as e:
                        logger.warning(f"Error accessing start date column for row {index}: {e}")
                
                lease_maturity_date = None
                if maturity_col is not None:
                    try:
                        maturity_date = row[maturity_col]
                        if not pd.isna(maturity_date):
                            if isinstance(maturity_date, str):
                                lease_maturity_date = maturity_date
                            else:
                                # Convert to string in YYYY-MM-DD format
                                try:
                                    lease_maturity_date = pd.to_datetime(maturity_date).strftime('%Y-%m-%d')
                                except Exception as e:
                                    logger.warning(f"Could not parse maturity date '{maturity_date}' for serial {serial_number}: {e}")
                    except Exception as e:
                        logger.warning(f"Error accessing maturity date column for row {index}: {e}")
                
                # Skip if both dates are None
                if lease_start_date is None and lease_maturity_date is None:
                    continue
                
                # Find asset by serial number
                asset = self.db.get_asset_by_serial(serial_number)
                if not asset:
                    not_found_count += 1
                    not_found_serials.append(serial_number)
                    continue
                
                # Update lease info
                success = self.db.update_lease_info(
                    asset['asset_id'],
                    lease_start_date=lease_start_date,
                    lease_maturity_date=lease_maturity_date
                )
                
                if success:
                    updated_count += 1
                else:
                    error_count += 1
                    
            except Exception as e:
                logger.error(f"Error processing row {index}: {e}")
                error_count += 1
        
        # Update expiry flags for all assets
        self.db.update_all_expiry_flags()
        
        # Log results
        logger.info(f"Processed {total_rows} rows with manual column selection")
        logger.info(f"Updated: {updated_count}, Not found: {not_found_count}, Errors: {error_count}")
        
        if not_found_count > 0:
            logger.info(f"Serials not found: {not_found_serials[:10]}...")
        
        return True, f"Processed {total_rows} rows. Updated: {updated_count}, Not found: {not_found_count}, Errors: {error_count}"

class ExcelSheetDialog:
    """Dialog to select a sheet and columns from an Excel workbook"""
    def __init__(self, parent, file_path):
        import pandas as pd
        import openpyxl
        
        self.result = None
        
        # Create dialog
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Select Excel Sheet and Columns")
        self.dialog.transient(parent)
        self.dialog.grab_set()
        
        # Make dialog resizable
        self.dialog.resizable(True, True)
        self.dialog.minsize(900, 600)
        
        # Center the dialog
        self.dialog.update_idletasks()
        width = self.dialog.winfo_width()
        height = self.dialog.winfo_height()
        x = (parent.winfo_width() - width) // 2 + parent.winfo_x()
        y = (parent.winfo_height() - height) // 2 + parent.winfo_y()
        self.dialog.geometry(f"+{x}+{y}")
        
        # Configure dialog grid
        self.dialog.columnconfigure(0, weight=1)
        self.dialog.rowconfigure(3, weight=1)  # Preview grid row
        
        # Instructions
        ttk.Label(self.dialog, text="Step 1: Select a worksheet from the Excel file", font=("", 12, "bold")).grid(
            row=0, column=0, padx=10, pady=(10, 5), sticky="w"
        )
        
        # Sheet selection
        sheet_frame = ttk.Frame(self.dialog)
        sheet_frame.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="ew")
        
        ttk.Label(sheet_frame, text="Worksheet:").pack(side=tk.LEFT, padx=(0, 5))
        
        # Get sheet names from the workbook
        self.workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_names = self.workbook.sheetnames
        
        # Store file path
        self.file_path = file_path
        
        # Sheet selection combobox
        self.sheet_var = tk.StringVar()
        sheet_combo = ttk.Combobox(sheet_frame, textvariable=self.sheet_var, values=sheet_names, state="readonly", width=40)
        sheet_combo.pack(side=tk.LEFT, padx=5)
        
        if sheet_names:
            sheet_combo.current(0)  # Select first sheet by default
        
        ttk.Button(sheet_frame, text="Load Preview", command=self.load_sheet_preview).pack(side=tk.LEFT, padx=5)
        
        # Column selection instructions
        ttk.Label(self.dialog, text="Step 2: Identify the columns in your worksheet:", font=("", 12, "bold")).grid(
            row=2, column=0, padx=10, pady=(10, 5), sticky="w"
        )
        
        # Preview frame
        self.preview_frame = ttk.LabelFrame(self.dialog, text="Data Preview")
        self.preview_frame.grid(row=3, column=0, padx=10, pady=10, sticky="nsew")
        
        # Configure preview frame
        self.preview_frame.columnconfigure(0, weight=1)
        self.preview_frame.rowconfigure(0, weight=1)
        
        # Create a Text widget for the preview
        self.preview_text = tk.Text(self.preview_frame, wrap=tk.NONE, height=15, width=90)
        self.preview_text.grid(row=0, column=0, sticky="nsew")
        
        # Add scrollbars
        y_scrollbar = ttk.Scrollbar(self.preview_frame, orient=tk.VERTICAL, command=self.preview_text.yview)
        y_scrollbar.grid(row=0, column=1, sticky="ns")
        x_scrollbar = ttk.Scrollbar(self.preview_frame, orient=tk.HORIZONTAL, command=self.preview_text.xview)
        x_scrollbar.grid(row=1, column=0, sticky="ew")
        
        self.preview_text.configure(yscrollcommand=y_scrollbar.set, xscrollcommand=x_scrollbar.set)
        
        # Column selection controls
        selection_frame = ttk.Frame(self.dialog)
        selection_frame.grid(row=4, column=0, padx=10, pady=10, sticky="ew")
        
        # Configure grid
        for i in range(3):
            selection_frame.columnconfigure(i, weight=1)
        
        # Serial Number column selection
        ttk.Label(selection_frame, text="Serial Number Column:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.serial_var = tk.StringVar()
        self.serial_combo = ttk.Combobox(selection_frame, textvariable=self.serial_var, state="readonly", width=30)
        self.serial_combo.grid(row=1, column=0, padx=5, pady=5, sticky="ew")
        
        # Lease Start Date column selection
        ttk.Label(selection_frame, text="Lease Start Date Column:").grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.start_var = tk.StringVar()
        self.start_combo = ttk.Combobox(selection_frame, textvariable=self.start_var, state="readonly", width=30)
        self.start_combo.grid(row=1, column=1, padx=5, pady=5, sticky="ew")
        
        # Lease Maturity Date column selection
        ttk.Label(selection_frame, text="Lease Maturity Date Column:").grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.maturity_var = tk.StringVar()
        self.maturity_combo = ttk.Combobox(selection_frame, textvariable=self.maturity_var, state="readonly", width=30)
        self.maturity_combo.grid(row=1, column=2, padx=5, pady=5, sticky="ew")
        
        # Sample data display (to help identify columns)
        self.sample_frame = ttk.LabelFrame(self.dialog, text="Column Sample Data")
        self.sample_frame.grid(row=5, column=0, padx=10, pady=10, sticky="ew")
        
        sample_grid = ttk.Frame(self.sample_frame)
        sample_grid.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Configure grid columns
        for i in range(3):
            sample_grid.columnconfigure(i, weight=1)
        
        # Labels for sample data
        self.serial_sample = ttk.Label(sample_grid, text="", wraplength=250)
        self.serial_sample.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        
        self.start_sample = ttk.Label(sample_grid, text="", wraplength=250)
        self.start_sample.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        
        self.maturity_sample = ttk.Label(sample_grid, text="", wraplength=250)
        self.maturity_sample.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        
        # Update sample data when column selections change
        self.serial_combo.bind("<<ComboboxSelected>>", self.update_sample_data)
        self.start_combo.bind("<<ComboboxSelected>>", self.update_sample_data)
        self.maturity_combo.bind("<<ComboboxSelected>>", self.update_sample_data)
        
        # Button frame
        button_frame = ttk.Frame(self.dialog)
        button_frame.grid(row=6, column=0, pady=10)
        
        def on_ok():
            sheet_name = self.sheet_var.get()
            serial_col = self.serial_var.get()
            start_col = self.start_var.get()
            maturity_col = self.maturity_var.get()
            
            if not sheet_name:
                messagebox.showerror("Error", "Please select a worksheet", parent=self.dialog)
                return
                
            if not serial_col:
                messagebox.showerror("Error", "Please select Serial Number column", parent=self.dialog)
                return
            
            # Start and maturity date columns are optional, but at least one should be provided
            if not start_col and not maturity_col:
                if not messagebox.askyesno("Warning", 
                                       "No date columns selected. Continue anyway?", 
                                       parent=self.dialog):
                    return
            
            self.result = {
                'sheet_name': sheet_name,
                'serial_col': serial_col,
                'start_col': start_col if start_col else None,
                'maturity_col': maturity_col if maturity_col else None
            }
            self.dialog.destroy()
        
        ttk.Button(button_frame, text="OK", command=on_ok).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Cancel", command=self.dialog.destroy).pack(side=tk.LEFT, padx=5)
        
        # Load preview of the first sheet
        if sheet_names:
            self.load_sheet_preview()
        
        # Wait for dialog to close
        parent.wait_window(self.dialog)
    
    def load_sheet_preview(self):
        """Load preview of the selected sheet"""
        import pandas as pd
        
        sheet_name = self.sheet_var.get()
        if not sheet_name:
            return
        
        try:
            # Read the selected sheet
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, nrows=15)
            
            # Clear preview text
            self.preview_text.config(state=tk.NORMAL)
            self.preview_text.delete(1.0, tk.END)
            
            # Format and display preview
            preview_text = f"Preview of sheet '{sheet_name}':\n\n"
            preview_text += df.to_string(index=True)
            self.preview_text.insert(tk.END, preview_text)
            self.preview_text.config(state=tk.DISABLED)
            
            # Update column selection dropdowns
            column_options = []
            
            # Add column letters (Excel style)
            for i, col in enumerate(df.columns):
                col_letter = self.get_column_letter(i)
                display_name = f"{col_letter}: {col}"
                column_options.append(display_name)
            
            # Update comboboxes
            self.serial_combo['values'] = column_options
            self.start_combo['values'] = column_options
            self.maturity_combo['values'] = column_options
            
            # Clear selections
            self.serial_combo.set('')
            self.start_combo.set('')
            self.maturity_combo.set('')
            
            # Try to auto-detect columns
            self.auto_detect_columns(df)
            
            # Store dataframe for sample data
            self.df = df
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load sheet: {str(e)}", parent=self.dialog)
            logger.error(f"Error loading Excel sheet: {e}")
    
    def get_column_letter(self, col_index):
        """Convert 0-based column index to Excel-style column letter"""
        result = ""
        while col_index >= 0:
            col_index, remainder = divmod(col_index, 26)
            result = chr(65 + remainder) + result
            col_index -= 1
        return result
    
    def auto_detect_columns(self, df):
        """Try to auto-detect relevant columns"""
        # Look for keywords in column names
        serial_keywords = ['serial', 'sn', 'device', 'asset']
        start_keywords = ['start', 'begin', 'lease start']
        maturity_keywords = ['end', 'maturity', 'expiry', 'expire', 'term']
        
        # For each column, check if any keyword is in the column name
        for i, col in enumerate(df.columns):
            col_letter = self.get_column_letter(i)
            col_str = str(col).lower()
            
            # Check for serial number column
            if any(kw in col_str for kw in serial_keywords):
                self.serial_combo.set(f"{col_letter}: {col}")
            
            # Check for lease start date column
            if any(kw in col_str for kw in start_keywords):
                self.start_combo.set(f"{col_letter}: {col}")
            
            # Check for lease maturity date column
            if any(kw in col_str for kw in maturity_keywords):
                self.maturity_combo.set(f"{col_letter}: {col}")
        
        # If no matches found, try excel standard columns for our expected data
        if not self.serial_var.get() and len(df.columns) > 0:
            self.serial_combo.current(0)  # First column (A) for serial
        
        if not self.start_var.get() and len(df.columns) > 8:
            self.start_combo.current(8)  # Column I for start date
            
        if not self.maturity_var.get() and len(df.columns) > 9:
            self.maturity_combo.current(9)  # Column J for maturity date
        
        # Update sample data
        self.update_sample_data()
    
    def update_sample_data(self, event=None):
        """Update sample data display based on selected columns"""
        if not hasattr(self, 'df') or self.df is None:
            return
            
        try:
            # Get selected columns
            serial_selection = self.serial_var.get()
            start_selection = self.start_var.get()
            maturity_selection = self.maturity_var.get()
            
            # Extract data samples
            if serial_selection:
                try:
                    col_name = serial_selection.split(": ", 1)[1] if ": " in serial_selection else serial_selection
                    sample_data = self.df[col_name].dropna().head(5).tolist()
                    self.serial_sample.config(text=f"Serial Number Samples:\n{', '.join(str(x) for x in sample_data)}")
                except Exception as e:
                    self.serial_sample.config(text=f"Error loading samples: {str(e)}")
            else:
                self.serial_sample.config(text="")
                
            if start_selection:
                try:
                    col_name = start_selection.split(": ", 1)[1] if ": " in start_selection else start_selection
                    sample_data = self.df[col_name].dropna().head(5).tolist()
                    self.start_sample.config(text=f"Start Date Samples:\n{', '.join(str(x) for x in sample_data)}")
                except Exception as e:
                    self.start_sample.config(text=f"Error loading samples: {str(e)}")
            else:
                self.start_sample.config(text="")
                
            if maturity_selection:
                try:
                    col_name = maturity_selection.split(": ", 1)[1] if ": " in maturity_selection else maturity_selection
                    sample_data = self.df[col_name].dropna().head(5).tolist()
                    self.maturity_sample.config(text=f"Maturity Date Samples:\n{', '.join(str(x) for x in sample_data)}")
                except Exception as e:
                    self.maturity_sample.config(text=f"Error loading samples: {str(e)}")
            else:
                self.maturity_sample.config(text="")
                
        except Exception as e:
            logger.error(f"Error updating sample data: {e}")
            # Don't re-raise the exception, just log it